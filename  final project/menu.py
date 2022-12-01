# 1: Загружать данные из файла должна уметь программа, сохранять, новые записи и производить поиск по существующим
# записям, сохраняет данные о человеке, а именно: ФИО, дата рождения, дата смерти (может отсутствовать) и пол.
# Программа должна уметь вычислять возраст человека (количество полных лет) на основании даты рождения и даты
# смерти или сегодняшней даты, если дата смерти отсутствует.
# 2: Поиск может производиться по имени, фамилии и отчеству и выдаёт все варианты, которые подходят под строку
# поиска (это может быть имя, или фамилия, или имя и фамилия, или только часть имени и т.д.)
# Дата рождения и дата смерти может вводиться в формате:
#
# 12.10.1980
# 11 10 2000
# 01/02/1995
# 3-9-2007.
from datetime import datetime
import openpyxl
import csv
import json
from openpyxl import load_workbook

from PyQt5 import QtCore, QtGui, QtWidgets


class Ui_MainWindow():
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(898, 556)
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        MainWindow.setFont(font)
        MainWindow.setStyleSheet("background-color: rgb(34, 34, 46);")
        MainWindow.setLocale(QtCore.QLocale(QtCore.QLocale.Ukrainian, QtCore.QLocale.Ukraine))
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.frame = QtWidgets.QFrame(self.centralwidget)
        self.frame.setGeometry(QtCore.QRect(0, 0, 331, 561))
        self.frame.setStyleSheet("background-color: rgb(251, 91, 93);")
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.label = QtWidgets.QLabel(self.frame)
        self.label.setGeometry(QtCore.QRect(70, 210, 201, 41))
        self.label.setStyleSheet("background-color: rgb(34, 34, 46);\n"
                                 "border: 2px solid #f66867;\n"
                                 "border-radius: 20\n"
                                 "\n"
                                 "\n"
                                 "")
        self.label.setObjectName("label")
        self.path_file = QtWidgets.QLineEdit(self.frame)
        self.path_file.setGeometry(QtCore.QRect(10, 260, 301, 41))
        self.path_file.setStyleSheet("font: 12pt \"MingLiU_HKSCS-ExtB\";\n"
                                     "background-color: rgb(34, 34, 46);\n"
                                     "border: 2px solid #f66867;\n"
                                     "border-radius: 20;\n"
                                     "color: white;")
        self.path_file.setText("")
        self.path_file.setAlignment(QtCore.Qt.AlignCenter)
        self.path_file.setObjectName("path_file")
        self.name = QtWidgets.QLineEdit(self.frame)
        self.name.setGeometry(QtCore.QRect(10, 360, 311, 41))
        self.name.setStyleSheet("font: 12pt \"MingLiU_HKSCS-ExtB\";\n"
                                "background-color: rgb(34, 34, 46);\n"
                                "border: 2px solid #f66867;\n"
                                "border-radius: 20;\n"
                                "color: white;")
        self.name.setText("")
        self.name.setAlignment(QtCore.Qt.AlignCenter)
        self.name.setObjectName("name")
        self.label_2 = QtWidgets.QLabel(self.frame)
        self.label_2.setGeometry(QtCore.QRect(10, 310, 311, 41))
        self.label_2.setStyleSheet("background-color: rgb(34, 34, 46);\n"
                                   "border: 2px solid #f66867;\n"
                                   "border-radius: 20\n"
                                   "\n"
                                   "")
        self.label_2.setObjectName("label_2")
        self.Search = QtWidgets.QPushButton(self.frame)
        self.Search.setGeometry(QtCore.QRect(30, 60, 261, 41))
        self.Search.setStyleSheet("font: 12pt \"MingLiU_HKSCS-ExtB\";\n"
                                  "background-color: rgb(34, 34, 34);\n"
                                  "border: 2px solid #f66867;\n"
                                  "color: white;")
        self.Search.setObjectName("Search")
        self.Add = QtWidgets.QPushButton(self.frame)
        self.Add.setGeometry(QtCore.QRect(30, 110, 261, 41))
        self.Add.setStyleSheet("font: 12pt \"MingLiU_HKSCS-ExtB\";\n"
                               "background-color: rgb(34, 34, 34);\n"
                               "border: 2px solid #f66867;\n"
                               "color: white;")
        self.Add.setObjectName("Add")
        self.Clear = QtWidgets.QPushButton(self.frame)
        self.Clear.setGeometry(QtCore.QRect(30, 160, 261, 41))
        self.Clear.setStyleSheet("font: 12pt \"MingLiU_HKSCS-ExtB\";\n"
                                 "background-color: rgb(34, 34, 34);\n"
                                 "border: 2px solid #f66867;\n"
                                 "color: white;")
        self.Clear.setObjectName("Clear")
        self.label_3 = QtWidgets.QLabel(self.frame)
        self.label_3.setGeometry(QtCore.QRect(10, 10, 311, 41))
        self.label_3.setStyleSheet("background-color: rgb(34, 34, 46);\n"
                                   "border: 2px solid #f66867;\n"
                                   "border-radius: 20\n"
                                   "\n"
                                   "")
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.frame)
        self.label_4.setGeometry(QtCore.QRect(60, 410, 211, 41))
        self.label_4.setStyleSheet("background-color: rgb(34, 34, 46);\n"
                                   "border: 2px solid #f66867;\n"
                                   "border-radius: 20\n"
                                   "\n"
                                   "")
        self.label_4.setObjectName("label_4")
        self.add2 = QtWidgets.QLineEdit(self.frame)
        self.add2.setGeometry(QtCore.QRect(10, 460, 311, 41))
        self.add2.setStyleSheet("font: 12pt \"MingLiU_HKSCS-ExtB\";\n"
                                  "background-color: rgb(34, 34, 46);\n"
                                  "border: 2px solid #f66867;\n"
                                  "border-radius: 20;\n"
                                  "color: white;")
        self.add2.setText("")
        self.add2.setAlignment(QtCore.Qt.AlignCenter)
        self.add2.setObjectName("name_2")
        self.result = QtWidgets.QLabel(self.centralwidget)
        self.result.setGeometry(QtCore.QRect(360, 110, 511, 421))
        self.result.setStyleSheet("font: 12pt \"MingLiU_HKSCS-ExtB\";\n"
                                  "background-color: rgb(34, 34, 46);\n"
                                  "border: 2px solid #f66867;\n"
                                  "border-radius: 30;\n"
                                  "color: white;\n"
                                  "")
        self.result.setLineWidth(10)
        self.result.setMidLineWidth(12)
        self.result.setText("")
        self.result.setAlignment(QtCore.Qt.AlignLeading | QtCore.Qt.AlignLeft | QtCore.Qt.AlignTop)
        self.result.setObjectName("result")
        self.lable_result = QtWidgets.QLabel(self.centralwidget)
        self.lable_result.setGeometry(QtCore.QRect(500, 40, 231, 41))
        self.lable_result.setStyleSheet("background-color: rgb(34, 34, 46);\n"
                                        "border: 2px solid #f66867;\n"
                                        "border-radius: 30\n"
                                        "\n"
                                        "")
        self.lable_result.setObjectName("lable_result")
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.add_functions()

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Searcher"))
        self.label.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" "
                                                    "font-size:16pt; color:#ffffff;\">Full path</span></p></body></html>"))
        self.label_2.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" "
                                                      "font-size:16pt; color:#ffffff;\">Name  surename midle name</span></p></body></html>"))
        self.Search.setText(_translate("MainWindow", "Search"))
        self.Add.setText(_translate("MainWindow", "Add"))
        self.Clear.setText(_translate("MainWindow", "Clear"))
        self.label_3.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" "
                                                      "font-size:16pt; color:#ffffff;\">Action menu</span></p></body></html>"))
        self.label_4.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" "
                                                      "font-size:16pt; color:#ffffff;\">For add</span></p></body></html>"))
        self.lable_result.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" "
                                                           "font-size:16pt; color:#ffffff;\">Result</span></p></body></html>"))

    def add_functions(self):
        self.Add.clicked.connect(lambda: self.add(self.add2.text(), self.path_file.text()))
        self.Clear.clicked.connect(lambda: self.clear())
        self.Search.clicked.connect(lambda: self.search(self.name.text(), self.path_file.text()))

    def clear(self):
        self.result.clear()
        self.name.clear()
        self.add2.clear()
    def add(self, info, file_path):
        if '.csv' in file_path:                                                            # csv файлы добавления
            info = info.split()
            if len(info) == 6:
                info = [f'{info[0]} {info[1]} {info[2]}', f'{info[3]}', f'{info[4]}', f'{info[5]}']
                with open(fr'{file_path}', 'a', encoding='utf-8') as sourses:
                    sourse = csv.writer(sourses)
                    sourse.writerow(info)
                    self.result.setText(self.result.text() + f'   {info}')
            elif len(info) > 6:
                self.result.setText('     \n \n \n \n     Ошибка, слишком много элементов')
            else:
                self.result.setText('     \n \n \n \n     Ошибка, слишком мало элементов')       #--------------
        elif '.txt' in file_path:                                                          # txt файлы добавления
            with open(fr'{file_path}', 'a') as sourses:
                sourse = sourses
                info = info.split()
                if len(info) == 6:
                    info = f"\n'{info[0]} {info[1]} {info[2]}','{info[3]}','{info[4]}','{info[5]}'"
                    sourse.write(info)
                    self.result.setText(f'   Был добавлен файл:\n  {info}')
                elif len(info) > 6:
                    self.result.setText('     \n \n \n \n     Ошибка, слишком много элементов')
                else: self.result.setText('     \n \n \n \n     Ошибка, не достаточно элементов') #--------------
        elif '.json' in file_path:      # json файлы добавления
            info = info.split(',')
            if len(info) == 4:
                q = 0
                with open(f'{file_path}', 'r') as source:
                    sources = json.load(source)
                    if len(sources) > q:
                        q = len(sources) + 1
                    info = {q:[f'{info[0]}', f'{info[1]}', f'{info[2]}', f'{info[3]}']}
                    sources.update(info)
                i = sources

                with open(fr'{file_path}', 'w') as source:
                    json.dump(i, source)
                self.result.setText(self.result.text() + f'  \n  Данные добавлены в файл:'
                                                         f'  \n  {info}')
            elif len(info) > 4: self.result.setText('     \n \n \n \n     Ошибка, слишком много значений')
            else: self.result.setText('     \n \n \n \n     Ошибка, слишком мало значений')       #--------------
        elif 'xlsx' in file_path:            # xlsx файлы добавления
            wb = load_workbook(file_path)
            ws = wb['Sheet1']
            info = info.split(',')
            print(info)
            if len(info) == 4:
                ws.append([f'{info[0]}', f'{info[1]}', f'{info[2]}', f'{info[3]}'])
                wb.save(file_path)
                wb.close()
                self.result.setText(f'   Был добавлен данные:\n  {info}')


            elif len(info) > 6:
                print('dwadaw')
            else:
                print('ad')  #

        else:
            self.result.setText('     \n \n \n \n     Ошибка, только файлы формата txt, csv или xlsx')


    def search(self, info, file_path):
        try:
            if '.csv' in file_path:                        # csv search
                try:
                    with open(fr'{file_path}', encoding='utf-8') as source:
                        reader = csv.reader(source)
                        for row in reader:
                            if info in row:
                                birth_day = row[2]
                                death_day = row[3]
                                if len(row[3]) > 1:
                                    birth_day = birth_day.split('.')
                                    death_day = death_day.split('.')
                                    death_year = death_day[2]
                                    birth_year = birth_day[2]
                                    age = int(death_year) - int(birth_year)
                                    self.result.setText(self.result.text() + f'    \n {*row, age}')
                                else:
                                    data = datetime.now()
                                    data = data.year
                                    birth_day = birth_day.split('.')
                                    birth_year = birth_day[2]
                                    age = data - int(birth_year)
                                    self.result.setText(self.result.text() + f'    \n {*row, age}')


                except FileNotFoundError:
                    self.result.setText('     \n         Ошибка: такого файла не существует') #--------------
            elif '.json' in file_path:                                                        # csv search
                try:
                    with open(f'{file_path}', 'r') as source:
                        sources = json.load(source)
                        for i in sources.values():
                            if info in i[0]:
                                if len(i[2]) > 1:
                                    birth_day = i[1]
                                    death_day = i[2]
                                    birth_day = birth_day.split('.')
                                    death_day = death_day.split('.')
                                    death_year = death_day[2]
                                    birth_year = birth_day[2]
                                    age = int(death_year) - int(birth_year)
                                    self.result.setText(self.result.text() + f'  ФИО Пол Дата родения Дата смерти:'
                                                                             f'  \n {i}'
                                                                             f' \n  Полные года: {age}')
                                else:
                                    birth_day = i[1]
                                    data = datetime.now()
                                    data = data.year
                                    birth_day = birth_day.split('.')
                                    birth_year = birth_day[2]
                                    age = data - int(birth_year)
                                    self.result.setText(self.result.text() + f'  ФИО Пол Дата родения Дата смерти:'
                                                                             f'  \n {i}'
                                                                             f' \n  Полные года: {age}')

                except FileNotFoundError:
                    self.result.setText('     \n          Ошибка: такого файла не существует') #--------------
            elif '.txt' in file_path:  # txt searcher
                d = False
                with open(f'{file_path}', 'r') as file_sourse:
                    base = list(file_sourse)
                    for i in base:
                        if info in i:
                            d = True
                            a = i
                            a = a.split("'")
                            birth_day = a[5]
                            death_day = a[7]
                            if len(death_day) > 1:
                                birth_year = birth_day.split('.')
                                birth_year = birth_year[2]
                                death_year = death_day.split('.')
                                death_year = death_year[2]
                                age = int(death_year) - int(birth_year)
                                self.result.setText(self.result.text() + f'\n  ФИО Пол Дата родения Дата смерти: '
                                                                         f'\n  {i} \n  Полные года: {age}')
                            else:
                                data = datetime.now()
                                data = data.year
                                birth_year = birth_day.split('.')
                                birth_year = birth_year[2]
                                age = data - int(birth_year)
                                self.result.setText(self.result.text() + f'\n  ФИО Пол Дата родения Дата смерти: '
                                                                         f'\n  {i} \n  Полные года: {age}')#--------------
                if d is False:
                    self.result.setText(f'     \n          пользователь {info} не был найден')
            elif '.xlsx' in file_path:                                                      # .xlsx
                book = openpyxl.open(fr'{file_path}')
                sheet = book.active
                for row in range(1, sheet.max_row + 1):
                    FIO = sheet[row][0].value
                    Birth_day = sheet[row][1].value
                    death_day = sheet[row][2].value
                    sex = sheet[row][3].value
                    if info in FIO:
                        date = datetime.now()
                        if len(death_day) > 1:
                            B = Birth_day.split('.')
                            B = B[2]
                            D = death_day.split('.')
                            D = D[2]

                            age = int(D) - int(B)
                            self.result.setText(self.result.text() + f'  \n  ФИО Дата родения Дата смерти Пол:'
                                                                     f'  \n  {FIO} {Birth_day} ,{death_day}, {sex}'
                                                                     f'  \n  Полные года: {age}')
                        else:
                            B = Birth_day.split('.')
                            B = B[2]
                            age = date.year - int(B)
                            self.result.setText(self.result.text() + f'  \n  ФИО Дата родения Дата смерти Пол:'
                                                                     f'  \n  {FIO} {Birth_day} ,{death_day}, {sex}'
                                                                     f'  \n  Полные года: {age}')
            else:
                self.result.setText('     \n     Ошибка, только файлы формата txt, csv или xlsx')#--------------
        except Exception: self.result.setText('     \n    Ошибка')




if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
